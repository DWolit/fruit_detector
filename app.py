from flask import Flask, request, jsonify, render_template, send_file
import cv2
import numpy as np
import base64
import json
import sqlite3
import io
from datetime import datetime
from ultralytics import YOLO
from openpyxl import Workbook

app = Flask(__name__)
model = YOLO('yolov8n.pt')

FRUIT_CLASSES = {46: 'банан', 47: 'яблоко', 49: 'апельсин'}
COLORS = {'яблоко': '#e74c3c', 'банан': '#f1c40f', 'апельсин': '#e67e22'}


def init_db():
    conn = sqlite3.connect('history.db')
    conn.execute('''CREATE TABLE IF NOT EXISTS results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp TEXT,
        filename TEXT,
        detections TEXT,
        total INTEGER
    )''')
    conn.commit()
    conn.close()


@app.route('/')
def index():
    return render_template('index.html', colors=json.dumps(COLORS))


@app.route('/process', methods=['POST'])
def process():
    if 'image' not in request.files:
        return jsonify({'error': 'Файл не загружен'}), 400

    file = request.files['image']
    img_bytes = file.read()
    img = cv2.imdecode(np.frombuffer(img_bytes, np.uint8), 1)

    if img is None:
        return jsonify({'error': 'Не удалось прочитать изображение'}), 400

    results = model(img)

    detections = []
    for r in results:
        for box in r.boxes:
            cls_id = int(box.cls[0])
            if cls_id in FRUIT_CLASSES:
                coords = box.xyxy[0]
                x1 = int(coords[0])
                y1 = int(coords[1])
                x2 = int(coords[2])
                y2 = int(coords[3])
                conf = round(float(box.conf[0]) * 100, 1)
                detections.append({
                    'x1': x1, 'y1': y1, 'x2': x2, 'y2': y2,
                    'label': FRUIT_CLASSES[cls_id],
                    'confidence': conf
                })

    print('Найдено объектов:', len(detections))

    _, buf = cv2.imencode('.jpg', img)
    img_b64 = base64.b64encode(buf).decode('utf-8')

    return jsonify({
        'image': img_b64,
        'detections': detections,
        'width': img.shape[1],
        'height': img.shape[0],
        'filename': file.filename
    })


@app.route('/save', methods=['POST'])
def save():
    data = request.get_json()
    detections = data.get('detections', [])
    filename = data.get('filename', 'unknown')

    conn = sqlite3.connect('history.db')
    conn.execute(
        'INSERT INTO results (timestamp, filename, detections, total) VALUES (?, ?, ?, ?)',
        (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), filename, json.dumps(detections), len(detections))
    )
    conn.commit()
    conn.close()

    return jsonify({'status': 'ok'})


@app.route('/delete/<int:record_id>', methods=['POST'])
def delete(record_id):
    conn = sqlite3.connect('history.db')
    conn.execute('DELETE FROM results WHERE id = ?', (record_id,))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})


@app.route('/history')
def history():
    conn = sqlite3.connect('history.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM results ORDER BY timestamp DESC')
    rows = cursor.fetchall()
    conn.close()

    records = []
    for row in rows:
        dets = json.loads(row[3])
        counts = {}
        for d in dets:
            label = d['label']
            if label in counts:
                counts[label] += 1
            else:
                counts[label] = 1
        records.append({
            'id': row[0],
            'timestamp': row[1],
            'filename': row[2],
            'detections': dets,
            'total': row[4],
            'counts': counts
        })

    return render_template('history.html', records=records)


@app.route('/export/<int:record_id>')
def export(record_id):
    conn = sqlite3.connect('history.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM results WHERE id = ?', (record_id,))
    row = cursor.fetchone()
    conn.close()

    if not row:
        return 'Запись не найдена', 404

    detections = json.loads(row[3])

    wb = Workbook()
    ws = wb.active
    ws.title = 'Результаты'

    ws['A1'] = 'Отчёт по подсчёту фруктов'
    ws['A2'] = 'Дата: ' + row[1]
    ws['A3'] = 'Файл: ' + row[2]
    ws['A4'] = 'Всего обнаружено: ' + str(row[4])
    ws.append([])
    ws.append(['№', 'Фрукт', 'Уверенность (%)'])

    for i, det in enumerate(detections, 1):
        ws.append([i, det['label'], det['confidence']])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name='fruits_' + str(record_id) + '.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
